# Importamos la clase Producto de la carpeta Entidades
from Entidades.Producto import Producto
# Importamos datetime para manejar fechas
from datetime import datetime
# Importamos openpyxl para trabajar con archivos Excel
import openpyxl
from openpyxl import Workbook

# Clase encargada de gestionar los productos
class ControladorProducto:
    def __init__(self):
        # Diccionario donde se almacenan los productos, con el código como clave
        self.productos = {}

    # Agrega un producto con la fecha actual generada automáticamente
    def agregar_producto(self, codigo, nombre, descripcion, proveedor, categoria, stock):
        # Verifica si ya existe un producto con el mismo código
        if codigo in self.productos:
            return False

        # Genera la fecha actual como cadena formateada
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Crea una instancia de Producto con los datos ingresados
        nuevo = Producto(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            fechaActualizacion=fecha_actual,
            proveedor=proveedor,
            categoria=categoria,
            stock=stock
        )
        # Agrega el producto al diccionario
        self.productos[codigo] = nuevo
        return True

    # Agrega un producto recibiendo una fecha personalizada (por ejemplo, desde Excel)
    def agregar_productoF(self, codigo, nombre, descripcion, proveedor, categoria, stock, fechaActualizacion):
        if codigo in self.productos:
            return False

        nuevo = Producto(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            proveedor=proveedor,
            categoria=categoria,
            stock=stock,
            fechaActualizacion=fechaActualizacion
        )
        self.productos[codigo] = nuevo
        return True

    # Actualiza el stock y fecha de actualización de un producto ya registrado
    def actualizar_stock(self, codigo, nuevo_stock):
        # Verifica que el producto exista
        if codigo not in self.productos:
            return False

        producto = self.productos[codigo]
        # Cambia el stock y actualiza la fecha
        producto.set_stock(nuevo_stock)
        producto.set_fechaActualizacion(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        return True

    # Devuelve la lista de productos (valores del diccionario)
    def listar_productos(self):
        return list(self.productos.values())

    # Devuelve un producto específico según su código
    def obtener_producto(self, codigo):
        return self.productos.get(codigo)

    # Carga productos desde un archivo de Excel (.xlsx)
    def importar_desde_excel(self, ruta_archivo):
        try:
            # Carga el archivo Excel
            wb = openpyxl.load_workbook(ruta_archivo)
            hoja = wb.active  # Obtiene la hoja activa

            # Recorre las filas a partir de la segunda (omitiendo encabezados)
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                if fila is None:
                    print("Fila vacía, se omitió.")
                    continue

                # Se extraen solo los primeros 7 valores (evita columnas extra)
                fila_limpia = fila[:7]

                # Validación de cantidad de campos y que no estén vacíos
                if len(fila_limpia) != 7 or any(campo is None for campo in fila_limpia):
                    print("Fila incompleta o con campos vacíos, se omitió:", fila)
                    continue

                # Se desempaquetan los valores
                tmpc, nombre, descripcion, proveedor, categoria, stock, fecha_actualizacion = fila_limpia

                # Verifica si el producto ya existe por código
                if str(tmpc) not in self.productos:
                    self.agregar_productoF(
                        str(tmpc),
                        nombre,
                        descripcion,
                        proveedor,
                        categoria,
                        int(stock),
                        fecha_actualizacion
                    )
                else:
                    print(f"Producto con código {tmpc} ya existe. Se omitió.")

            return True

        except Exception as e:
            print("Error al cargar productos desde Excel:", e)
            return False

    # Exporta todos los productos a un archivo Excel
    def exportar_a_excel(self, ruta_archivo):
        try:
            wb = Workbook()            # Crea un nuevo libro Excel
            hoja = wb.active           # Obtiene la hoja activa
            hoja.title = "Productos"   # Le asigna un nombre a la hoja

            # Escribe la fila de encabezados
            hoja.append(["Código", "Nombre", "Descripción", "Proveedor", "Categoría", "Stock", "Fecha Actualización"])

            # Escribe una fila por cada producto
            for producto in self.productos.values():
                hoja.append([
                    producto.get_codigo(),
                    producto.get_nombre(),
                    producto.get_descripcion(),
                    producto.get_proveedor(),
                    producto.get_categoria(),
                    producto.get_stock(),
                    producto.get_fechaActualizacion()
                ])

            wb.save(ruta_archivo)  # Guarda el archivo en la ruta indicada
            return True
        except Exception as e:
            print("Error al exportar productos a Excel:", e)
            return False
